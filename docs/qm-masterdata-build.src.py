from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAVY="12324F"; TEAL="147A6E"; AMBER="FFF3E0"; GREY="F5F8F8"; RED="FDF0F0"
F=lambda **k: Font(name="Arial", **k)
hdr=F(bold=True,color="FFFFFF",size=9); body=F(size=9); bodyb=F(bold=True,size=9)
mono=Font(name="Courier New",size=9,bold=True,color=NAVY)
fill_hdr=PatternFill("solid",fgColor=NAVY); fill_sec=PatternFill("solid",fgColor=TEAL)
fill_in=PatternFill("solid",fgColor=AMBER); fill_alt=PatternFill("solid",fgColor=GREY)
thin=Side(style="thin",color="D6DCE4"); bd=Border(bottom=thin)
wrap=Alignment(wrap_text=True,vertical="top"); top=Alignment(vertical="top")

wb=Workbook()

# ---------------- MIC MASTER ----------------
MIC=[
# stage, mic, short text, type, dec, uom, target, lsl, usl, catalog, selset, tier, standard, sampling, note
("1","S1_DEN","Linear density - denier verification","Quantitative",1,"DTEX","nominal","-2.0%","+2.0%","","","1","ISO 2060 / ASTM D1907","20 breaks / 5 pkgs",""),
("1","S1_ELG","Elongation at break","Quantitative",1,"%","23.0","18.0","28.0","","","1","ISO 2062 / ASTM D2256","20 breaks / 5 pkgs",""),
("1","S1_BWS","Boiling water shrinkage","Quantitative",1,"%","7.0","5.0","9.0","","","1","ASTM D2259","5 packages","Also enforce range <= 1.5% across the 5"),
("1","S1_BFL","Broken filaments","Quantitative",0,"EA","0","","2","","","1","Visual under tension","ISO 2859-1 GIL II",""),
("1","S1_PKD","Package density","Quantitative",2,"G/CM3","0.39","0.36","0.42","","","1","Mass / volume","ISO 2859-1 GIL II",""),
("1","S1_PKW","Package weight","Quantitative",2,"KG","nominal","-3.0%","+3.0%","","","1","Weighing","ISO 2859-1 GIL II",""),
("1","S1_GRD","Grade","Qualitative","","","min AA","","","1","ZQC_GRADE","1","Internal grading rules","ISO 2859-1 GIL II","BLOCKED - grade code list required (Q2)"),
("1","S1_IDT","Lot identity vs SAP batch","Qualitative","","","Pass","","","1","ZQC_PF","1","Tag vs GREY_CODE","100% of vendor batches","Auto-fail if batch draws >1 vendor lot"),
("2","S2_DE","Shade vs standard - delta E CMC(2:1)","Quantitative",2,"","0.00","","1.00","","","1","ISO 105-J03 / AATCC EP6","3 pkgs: top/mid/bottom",""),
("2","S2_INOUT","Inside-outside package variation","Qualitative","","","min 4-5","","","1","ZQC_GREY","1","ISO 105-A02","2 pkgs x 2 layers",""),
("2","S2_P2P","Package-to-package within batch","Qualitative","","","min 4","","","1","ZQC_GREY","1","ISO 105-A02","3 pkgs: top/mid/bottom",""),
("2","S2_B2B","Batch-to-batch vs master","Qualitative","","","min 4","","","1","ZQC_GREY","1","ISO 105-A02","1 composite",""),
("2","S2_LEVEL","Levelness / barre","Qualitative","","","min 4","","","1","ZQC_GREY","1","ISO 105-A02","Knitted sock, D65",""),
("2","S2_STAIN","Stains, tip-dyeing, rust","Qualitative","","","Pass","","","1","ZQC_PF","1","Visual","ISO 2859-1 GIL II","Defect code required on fail"),
("2","S2_WFC","Washing fastness - colour change","Qualitative","","","min 4","","","1","ZQC_GREY","2","ISO 105-C06 A1S","1 composite / recipe",""),
("2","S2_WFS","Washing fastness - staining","Qualitative","","","min 3-4","","","1","ZQC_GREY","2","ISO 105-C06 A1S","1 composite / recipe",""),
("2","S2_RUBD","Rubbing fastness - dry","Qualitative","","","min 4","","","1","ZQC_GREY","2","ISO 105-X12","1 composite / recipe",""),
("2","S2_RUBW","Rubbing fastness - wet","Qualitative","","","min 3","","","1","ZQC_GREY","2","ISO 105-X12","1 composite / recipe",""),
("2","S2_PSPA","Perspiration fastness - acid","Qualitative","","","min 4","","","1","ZQC_GREY","2","ISO 105-E04","1 composite / recipe",""),
("2","S2_PSPB","Perspiration fastness - alkaline","Qualitative","","","min 4","","","1","ZQC_GREY","2","ISO 105-E04","1 composite / recipe",""),
("2","S2_SUBL","Sublimation fastness - dry heat","Qualitative","","","min 4","","","1","ZQC_GREY","2","ISO 105-P01","1 composite / recipe","180C / 30s"),
("2","S2_HCHO","Free formaldehyde","Quantitative",0,"PPM","0","","75","","","3","ISO 14184-1","1 / certification cycle","External cert - capture ref no."),
("2","S2_AZO","Banned azo amines","Quantitative",0,"PPM","0","","20","","","3","EN ISO 14362-1","1 / certification cycle","External cert - capture ref no."),
("3","S3_PKW","Package weight - net","Quantitative",2,"KG","nominal","-1.5%","+1.5%","","","1","Weighing","ISO 2859-1 GIL II",""),
("3","S3_PKL","Package length","Quantitative",0,"M","nominal","-2.0%","+2.0%","","","2","ISO 2060","Winder counter + verify",""),
("3","S3_PKD","Package density","Quantitative",2,"G/CM3","0.46","0.42","0.50","","","1","Mass / volume","ISO 2859-1 GIL II",""),
("3","S3_TAPER","Cone taper and traverse","Qualitative","","","Pass","","","1","ZQC_PF","1","Gauge vs cone format","ISO 2859-1 GIL II",""),
("3","S3_RIBB","Ribboning / patterning","Qualitative","","","Pass","","","1","ZQC_PF","1","Visual","ISO 2859-1 GIL II",""),
("3","S3_NOSE","Nose bulging / soft-hard edge","Qualitative","","","Pass","","","1","ZQC_PF","1","Visual","ISO 2859-1 GIL II",""),
("3","S3_BFL","Broken filaments","Quantitative",0,"EA","0","","2","","","1","Visual under tension","ISO 2859-1 GIL II","Compare vs S1_BFL - delta is process-added"),
("3","S3_TENS","Winding tension","Quantitative",1,"CN","target","-10.0%","+10.0%","","","2","Tensiometer","Per winding position","Also CV% <= 6"),
("3","S3_DEN","Final denier","Quantitative",1,"DTEX","nominal","-2.0%","+2.0%","","","2","ISO 2060","20 breaks","Paired with S1_DEN"),
("3","S3_ELG","Final elongation","Quantitative",1,"%","23.0","18.0","28.0","","","2","ISO 2062","20 breaks","Loss vs S1_ELG <= 8%"),
("3","S3_CONT","Contamination / stain","Qualitative","","","Pass","","","1","ZQC_PF","1","Visual","ISO 2859-1 GIL II","Defect code required on fail"),
("3","S3_MOIS","Moisture","Quantitative",2,"%","0.30","","0.50","","","1","ISO 6741-1","1 composite",""),
("3","S3_OIL","Oil percent - coning oil","Quantitative",2,"%","1.50","1.00","2.00","","","1","Solvent extraction","1 composite","TARGET IS A PLACEHOLDER - confirm applicator setting"),
]

ws=wb.active; ws.title="2 MIC Master"
cols=[("Stage",7),("MIC",11),("Short text (QS21, max 40)",40),("Char. type",13),("Dec.",6),("UoM",8),
("Target",11),("Lower limit",12),("Upper limit",12),("Cat. type",10),("Selected set",13),
("Tier",6),("Standard / method",26),("Sampling",24),("Note / open item",42)]
for i,(t,w) in enumerate(cols,1):
    c=ws.cell(row=1,column=i,value=t); c.font=hdr; c.fill=fill_hdr; c.alignment=wrap
    ws.column_dimensions[get_column_letter(i)].width=w
ws.row_dimensions[1].height=30
for r,row in enumerate(MIC,2):
    for i,v in enumerate(row,1):
        c=ws.cell(row=r,column=i,value=v); c.font=mono if i==2 else body
        c.alignment=top if i not in (3,13,14,15) else wrap; c.border=bd
        if r%2==0: c.fill=fill_alt
    if "BLOCKED" in row[14] or "PLACEHOLDER" in row[14]:
        for i in range(1,16): ws.cell(row=r,column=i).fill=PatternFill("solid",fgColor=AMBER)
ws.freeze_panes="C2"
ws.auto_filter.ref=f"A1:O{len(MIC)+1}"

# ---------------- README ----------------
rd=wb.create_sheet("1 README",0)
rd.column_dimensions["A"].width=3; rd.column_dimensions["B"].width=42; rd.column_dimensions["C"].width=16
rd.column_dimensions["D"].width=62
def sec(r,t):
    c=rd.cell(row=r,column=2,value=t); c.font=F(bold=True,color="FFFFFF",size=10); c.fill=fill_sec
    for k in (3,4): rd.cell(row=r,column=k).fill=fill_sec
def kv(r,k,v,f=None):
    a=rd.cell(row=r,column=2,value=k); a.font=bodyb
    b=rd.cell(row=r,column=3,value=v); b.font=body
    if f: b.fill=f
    d=rd.cell(row=r,column=4,value="" ); d.font=body
def txt(r,t):
    c=rd.cell(row=r,column=2,value=t); c.font=body; c.alignment=wrap
    rd.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)

t=rd.cell(row=1,column=2,value="KGPL Yarn Dyeing QC - QM master data build sheet")
t.font=F(bold=True,size=15,color=NAVY)
rd.cell(row=2,column=2,value="Companion to Specification v2.0. Everything needed to create the QM master data in KSD client 500.").font=F(size=9,italic=True,color="5A6472")

sec(4,"Contents")
for i,(n,d) in enumerate([
 ("2 MIC Master","36 master inspection characteristics for QS21"),
 ("3 Code Groups","Code groups and codes for QS41"),
 ("4 Selected Sets","Selected sets with valuation for QS51"),
 ("5 Inspection Plans","Characteristic assignment per stage for QP01"),
 ("6 Sampling","Sampling procedures for QDV1 and the ISO 2859-1 table"),
 ("7 Usage Decisions","UD code group for catalog 3"),
 ("8 Open Items","What blocks entry, and what needs your numbers")],5):
    rd.cell(row=i,column=2,value=n).font=bodyb
    rd.cell(row=i,column=3,value="").font=body
    c=rd.cell(row=i,column=4,value=d); c.font=body

sec(13,"Counts - live formulas over sheet '2 MIC Master'")
rows=[("Total characteristics","=COUNTA('2 MIC Master'!B2:B37)"),
("Stage 1 - raw material","=COUNTIF('2 MIC Master'!A2:A37,\"1\")"),
("Stage 2 - after dyeing","=COUNTIF('2 MIC Master'!A2:A37,\"2\")"),
("Stage 3 - after winding","=COUNTIF('2 MIC Master'!A2:A37,\"3\")"),
("Quantitative","=COUNTIF('2 MIC Master'!D2:D37,\"Quantitative\")"),
("Qualitative","=COUNTIF('2 MIC Master'!D2:D37,\"Qualitative\")"),
("Tier 1 - every batch / lot","=COUNTIF('2 MIC Master'!L2:L37,\"1\")"),
("Tier 2 - per recipe / periodic","=COUNTIF('2 MIC Master'!L2:L37,\"2\")"),
("Tier 3 - certification cycle","=COUNTIF('2 MIC Master'!L2:L37,\"3\")")]
for i,(k,f) in enumerate(rows,14):
    rd.cell(row=i,column=2,value=k).font=body
    c=rd.cell(row=i,column=3,value=f); c.font=bodyb; c.alignment=Alignment(horizontal="left")

sec(24,"Legend")
rd.cell(row=25,column=2,value="Amber row").font=bodyb
rd.cell(row=25,column=3,value="").fill=fill_in
rd.cell(row=25,column=4,value="Blocked or placeholder - do not enter into SAP until resolved. See sheet 8.").font=body
rd.cell(row=26,column=2,value="'nominal' / 'target'").font=bodyb
rd.cell(row=26,column=4,value="Limit is relative to the material's nominal value; enter the absolute figure per material in QP01, not in QS21.").font=body
rd.cell(row=27,column=2,value="'min 4-5'").font=bodyb
rd.cell(row=27,column=4,value="Qualitative threshold - set by which codes are valuated as accepted in the selected set (sheet 4), not by a spec limit.").font=body

sec(29,"Correction to Specification v2.0")
txt(30,"v2.0 states 37 characteristics. The correct figure is 36 - S1_GRD was counted twice, once inside the stage 1 table and once again in the master data list. This workbook is authoritative: 8 + 15 + 13 = 36.")

sec(33,"Before you start")
for i,l in enumerate([
 "1. Resolve sheet 8 first. Two characteristics cannot be created until you supply a code list and a target.",
 "2. Check the units of measure exist in CUNI. DTEX, CN and G/CM3 are not in the SAP standard set and may need creating.",
 "3. Create code groups and selected sets (sheets 3 and 4) BEFORE the qualitative MICs - QS21 will reject a selected set that does not exist.",
 "4. Create MICs at plant level with status 'Released' (2), then assign to inspection plans in QP01.",
 "5. Percentage limits ('-2.0%') are relative. Enter them as absolute values per material in the inspection plan."],34):
    txt(i,l)

# ---------------- 3 CODE GROUPS ----------------
cg=wb.create_sheet("3 Code Groups")
CG=[("ZQC_GREY","1","Grey scale rating ISO 105-A02 / A03",[
 ("GS10","1 - severe change"),("GS15","1-2"),("GS20","2"),("GS25","2-3"),("GS30","3"),
 ("GS35","3-4"),("GS40","4"),("GS45","4-5"),("GS50","5 - no change")]),
("ZQC_PF","1","Pass / fail attribute",[("PASS","Pass"),("FAIL","Fail")]),
("ZQC_GRADE","1","Yarn grade - BLOCKED, see sheet 8",[("AA","Grade AA"),("A","Grade A"),("B","Grade B")]),
("ZQC_DEF2","9","Defect codes - after dyeing",[
 ("D201","Barre / streaky"),("D202","Tip dyeing"),("D203","Oil stain"),("D204","Rust stain"),
 ("D205","Inside-outside variation"),("D206","Shade off - lighter"),("D207","Shade off - darker"),
 ("D208","Dye spot / speck"),("D209","Crushed package")]),
("ZQC_DEF3","9","Defect codes - after winding",[
 ("D301","Ribboning / patterning"),("D302","Nose bulge"),("D303","Soft edge"),("D304","Hard edge"),
 ("D305","Damaged cone"),("D306","Oil stain from winder"),("D307","Foreign fibre"),
 ("D308","Broken filaments high"),("D309","Wrong label")])]
cg.column_dimensions["A"].width=14; cg.column_dimensions["B"].width=12
cg.column_dimensions["C"].width=40; cg.column_dimensions["D"].width=12; cg.column_dimensions["E"].width=34
for i,t in enumerate(["Code group","Catalog type","Group description","Code","Code description"],1):
    c=cg.cell(row=1,column=i,value=t); c.font=hdr; c.fill=fill_hdr
r=2
for grp,cat,desc,codes in CG:
    for j,(code,cd) in enumerate(codes):
        cg.cell(row=r,column=1,value=grp if j==0 else "").font=mono if j==0 else body
        cg.cell(row=r,column=2,value=cat if j==0 else "").font=body
        cg.cell(row=r,column=3,value=desc if j==0 else "").font=body
        cg.cell(row=r,column=4,value=code).font=mono
        cg.cell(row=r,column=5,value=cd).font=body
        for k in range(1,6):
            cg.cell(row=r,column=k).border=bd
            if grp=="ZQC_GRADE": cg.cell(row=r,column=k).fill=fill_in
        r+=1
cg.freeze_panes="A2"

# ---------------- 4 SELECTED SETS ----------------
ss=wb.create_sheet("4 Selected Sets")
ss.column_dimensions["A"].width=14; ss.column_dimensions["B"].width=12; ss.column_dimensions["C"].width=34
ss.column_dimensions["D"].width=10; ss.column_dimensions["E"].width=13; ss.column_dimensions["F"].width=46
for i,t in enumerate(["Selected set","Catalog type","Description","Code","Valuation","Applies to / note"],1):
    c=ss.cell(row=1,column=i,value=t); c.font=hdr; c.fill=fill_hdr
SS=[("ZQC_GREY","1","Grey scale - threshold set per characteristic in QP01",
  [("GS10","Reject"),("GS15","Reject"),("GS20","Reject"),("GS25","Reject"),("GS30","Reject"),
   ("GS35","see note"),("GS40","Accept"),("GS45","Accept"),("GS50","Accept")],
  "Default shown is for a 'min 4' characteristic. For 'min 3' (S2_RUBW) GS30 and GS35 also accept; for 'min 4-5' (S2_INOUT) GS40 rejects. Create three variants rather than one - see note below."),
 ("ZQC_PF","1","Pass / fail",[("PASS","Accept"),("FAIL","Reject")],"All attribute characteristics"),
 ("ZQC_GRADE","1","Yarn grade",[("AA","Accept"),("A","Accept"),("B","Reject")],"BLOCKED - confirm whether grade B is acceptable for any product")]
r=2
for st,cat,desc,codes,note in SS:
    for j,(code,val) in enumerate(codes):
        ss.cell(row=r,column=1,value=st if j==0 else "").font=mono if j==0 else body
        ss.cell(row=r,column=2,value=cat if j==0 else "").font=body
        ss.cell(row=r,column=3,value=desc if j==0 else "").font=body
        ss.cell(row=r,column=4,value=code).font=mono
        c=ss.cell(row=r,column=5,value=val); c.font=bodyb
        if val=="Reject": c.fill=PatternFill("solid",fgColor=RED)
        elif val=="Accept": c.fill=PatternFill("solid",fgColor="F1F9F2")
        ss.cell(row=r,column=6,value=note if j==0 else "").font=body
        ss.cell(row=r,column=6).alignment=wrap
        for k in range(1,7): ss.cell(row=r,column=k).border=bd
        if st=="ZQC_GRADE":
            for k in range(1,7): ss.cell(row=r,column=k).fill=fill_in
        r+=1
r+=1
c=ss.cell(row=r,column=1,value="Three grey-scale variants are required, because the acceptance threshold differs by characteristic:")
c.font=bodyb; ss.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
for i,(nm,thr,chars) in enumerate([
 ("ZQC_GREY3","accept GS30 and above (min 3)","S2_RUBW"),
 ("ZQC_GREY4","accept GS40 and above (min 4)","S2_P2P, S2_B2B, S2_LEVEL, S2_WFC, S2_RUBD, S2_PSPA, S2_PSPB, S2_SUBL"),
 ("ZQC_GREY45","accept GS45 and above (min 4-5)","S2_INOUT")],r+1):
    ss.cell(row=i,column=1,value=nm).font=mono
    ss.cell(row=i,column=2,value=thr).font=body
    ss.merge_cells(start_row=i,start_column=2,end_row=i,end_column=3)
    ss.cell(row=i,column=4,value=chars).font=body
    ss.merge_cells(start_row=i,start_column=4,end_row=i,end_column=6)
    ss.cell(row=i,column=4).alignment=wrap
ss.cell(row=r+4,column=1,value="S2_WFS (min 3-4) needs a fourth variant ZQC_GREY35 accepting GS35 and above.").font=F(size=9,italic=True)
ss.merge_cells(start_row=r+4,start_column=1,end_row=r+4,end_column=6)

# ---------------- 5 INSPECTION PLANS ----------------
ip=wb.create_sheet("5 Inspection Plans")
for i,(t,w) in enumerate([("Plan / stage",22),("Insp. type",11),("Op.",6),("Seq.",6),("MIC",11),
  ("Short text",38),("Sampling proc.",17),("Insp. scope",15),("Tier",6),("Notes",40)],1):
    c=ip.cell(row=1,column=i,value=t); c.font=hdr; c.fill=fill_hdr; c.alignment=wrap
    ip.column_dimensions[get_column_letter(i)].width=w
PLAN={"1":("ZQC_PLAN_S1 - Raw material","01","0010"),
      "2":("ZQC_PLAN_S2 - After dyeing","03","0010"),
      "3":("ZQC_PLAN_S3 - After winding","04","0010")}
r=2; last=None
for stage,mic,txt_,typ,dec,uom,tgt,lsl,usl,cat,sel,tier,std,samp,note in MIC:
    pl,ityp,op=PLAN[stage]
    ip.cell(row=r,column=1,value=pl if stage!=last else "").font=bodyb if stage!=last else body
    ip.cell(row=r,column=2,value=ityp if stage!=last else "").font=body
    ip.cell(row=r,column=3,value=op if stage!=last else "").font=body
    ip.cell(row=r,column=4,value=f"{(r-1)*10:04d}").font=body
    ip.cell(row=r,column=5,value=mic).font=mono
    ip.cell(row=r,column=6,value=txt_).font=body; ip.cell(row=r,column=6).alignment=wrap
    sp = "ZQC_ATTR" if "2859" in samp else ("ZQC_FIXED" if samp else "ZQC_100")
    ip.cell(row=r,column=7,value=sp).font=body
    ip.cell(row=r,column=8,value=samp).font=body; ip.cell(row=r,column=8).alignment=wrap
    ip.cell(row=r,column=9,value=tier).font=bodyb
    ip.cell(row=r,column=10,value=note).font=body; ip.cell(row=r,column=10).alignment=wrap
    for k in range(1,11):
        ip.cell(row=r,column=k).border=bd
        if r%2==0: ip.cell(row=r,column=k).fill=fill_alt
    last=stage; r+=1
ip.freeze_panes="E2"; ip.auto_filter.ref=f"A1:J{r-1}"

# ---------------- 6 SAMPLING ----------------
sm=wb.create_sheet("6 Sampling")
sm.cell(row=1,column=1,value="Sampling procedures for QDV1").font=F(bold=True,size=12,color=NAVY)
for i,(t,w) in enumerate([("Procedure",14),("Type",22),("Valuation",22),("Basis",34),("Used by",46)],1):
    c=sm.cell(row=3,column=i,value=t); c.font=hdr; c.fill=fill_hdr; c.alignment=wrap
    sm.column_dimensions[get_column_letter(i)].width=w
SMP=[("ZQC_ATTR","Attribute, single sampling","Acceptance number","ISO 2859-1, GIL II, normal severity","All pass/fail and defect-count characteristics"),
 ("ZQC_FIXED","Fixed sample size","Manual / spec limits","Fixed n per the Sampling column of sheet 2","Quantitative laboratory tests"),
 ("ZQC_100","100 percent","Manual","Every unit","S1_IDT only")]
for i,row in enumerate(SMP,4):
    for j,v in enumerate(row,1):
        c=sm.cell(row=i,column=j,value=v); c.font=mono if j==1 else body; c.alignment=wrap; c.border=bd
sm.cell(row=9,column=1,value="ISO 2859-1 - General Inspection Level II, normal severity").font=F(bold=True,size=11,color=NAVY)
for i,(t,w) in enumerate([("Lot size (packages)",22),("Code letter",14),("Sample size",14),
   ("Ac / Re at AQL 1.0",22),("Ac / Re at AQL 2.5",22),("Ac / Re at AQL 4.0",22)],1):
    c=sm.cell(row=11,column=i,value=t); c.font=hdr; c.fill=fill_hdr; c.alignment=wrap
    if w>sm.column_dimensions[get_column_letter(i)].width: sm.column_dimensions[get_column_letter(i)].width=w
TBL=[("2 - 8","A",2,"0 / 1","0 / 1","0 / 1"),("9 - 15","B",3,"0 / 1","0 / 1","0 / 1"),
 ("16 - 25","C",5,"0 / 1","0 / 1","0 / 1"),("26 - 50","D",8,"0 / 1","0 / 1","1 / 2"),
 ("51 - 90","E",13,"0 / 1","1 / 2","1 / 2"),("91 - 150","F",20,"1 / 2","1 / 2","2 / 3"),
 ("151 - 280","G",32,"1 / 2","2 / 3","3 / 4"),("281 - 500","H",50,"1 / 2","3 / 4","5 / 6"),
 ("501 - 1200","J",80,"2 / 3","5 / 6","7 / 8")]
for i,row in enumerate(TBL,12):
    for j,v in enumerate(row,1):
        c=sm.cell(row=i,column=j,value=v); c.font=body; c.border=bd
        if i%2==0: c.fill=fill_alt
sm.cell(row=22,column=1,value="AQL assignment").font=F(bold=True,size=10,color=TEAL)
for i,(a,d) in enumerate([("AQL 1.0 - critical","S1_IDT, S2_HCHO, S2_AZO - identity and chemical compliance"),
 ("AQL 2.5 - major","Shade, package integrity, broken filaments, contamination"),
 ("AQL 4.0 - minor","Cosmetic only")],23):
    sm.cell(row=i,column=1,value=a).font=bodyb
    c=sm.cell(row=i,column=2,value=d); c.font=body
    sm.merge_cells(start_row=i,start_column=2,end_row=i,end_column=6)
sm.cell(row=27,column=1,value="Lot size is the cheese count (ZPP_BATCHN.CHEESES) or the cone count from the packing record, so the app derives the sample size itself.").font=F(size=9,italic=True)
sm.merge_cells(start_row=27,start_column=1,end_row=27,end_column=6)

# ---------------- 7 USAGE DECISIONS ----------------
ud=wb.create_sheet("7 Usage Decisions")
for i,(t,w) in enumerate([("Code",9),("Description",26),("Trigger",44),("Stock / process consequence",44),("Follow-on",30)],1):
    c=ud.cell(row=1,column=i,value=t); c.font=hdr; c.fill=fill_hdr; c.alignment=wrap
    ud.column_dimensions[get_column_letter(i)].width=w
UD=[("A","Accept","All tier-1 characteristics within limits; recipe holds a valid tier-2 qualification","Unrestricted use; released to next stage",""),
 ("A1","Accept with deviation","Marginal result; concession signed by QC Manager","Unrestricted; deviation note carried on the batch","Requires approver authorisation"),
 ("RD","Re-dye / correct","delta E above 1.0 but shade correctable; unlevelness at grade 3","Blocked; batch reopened for correction","WIP Batch Close (reason QC-REDYE) then Cancel Production Confirmation"),
 ("ST","Strip and re-dye","Shade not correctable by topping; severe barre","Blocked; stripping process","Mandatory S3_ELG re-check after"),
 ("DG","Downgrade","Cosmetic or minor faults; yarn sound","Restricted use; B-grade stock",""),
 ("R","Reject","Chemical compliance failure; identity failure; elongation loss above 8%","Blocked; scrap or supplier return","")]
for i,row in enumerate(UD,2):
    for j,v in enumerate(row,1):
        c=ud.cell(row=i,column=j,value=v); c.font=mono if j==1 else body; c.alignment=wrap; c.border=bd
        if i%2==0: c.fill=fill_alt
ud.cell(row=9,column=1,value="Code group ZQC_UD, catalog type 3. Two corrections maximum before a QC Manager signature is required for a third.").font=F(size=9,italic=True)
ud.merge_cells(start_row=9,start_column=1,end_row=9,end_column=5)

# ---------------- 8 OPEN ITEMS ----------------
oi=wb.create_sheet("8 Open Items")
for i,(t,w) in enumerate([("#",5),("Blocks",16),("Item",34),("What is needed",50),("Your answer",34)],1):
    c=oi.cell(row=1,column=i,value=t); c.font=hdr; c.fill=fill_hdr; c.alignment=wrap
    oi.column_dimensions[get_column_letter(i)].width=w
OI=[("1","S1_GRD","Grade code list","Is Grade supplier-declared or QC-assigned? If QC-assigned, the grade definitions and which grades are acceptable. Sheet 3 assumes AA / A / B with B rejected.",""),
 ("2","S3_OIL","Coning oil target","The 1.00-2.00% in sheet 2 is a placeholder. Replace with your winder applicator setting and tolerance.",""),
 ("3","Units","DTEX, CN, G/CM3","These units are not in the SAP standard set. Confirm they exist in CUNI or have Basis create them.",""),
 ("4","S2_HCHO / S2_AZO","External or in-house","If external, the app captures a certificate reference rather than a live measurement, and the screen design changes.",""),
 ("5","Stage 1 frequency","Deliveries per day","Stage 1 is per supplier delivery lot, not per dye batch. The number drives the daily volume estimate.",""),
 ("6","All limits","Process capability","Every limit is an industry-typical proposal. Check each against your own capability - a limit the plant cannot meet teaches operators to ignore the system.",""),
 ("7","S1_BWS","Range limit","Spec v2 adds range <= 1.5% across 5 packages as a partial substitute for the dye uniformity test that was cut. Confirm this is workable.",""),
 ("8","Recipe ID","Tier 2 automation","Is the dye recipe separately identified in SAP, or only DYE_CODE? Tier 2 skipping depends on knowing the first batch of a recipe.","")]
for i,row in enumerate(OI,2):
    for j,v in enumerate(row,1):
        c=oi.cell(row=i,column=j,value=v); c.font=mono if j==2 else body; c.alignment=wrap; c.border=bd
        if j==5: c.fill=fill_in
oi.freeze_panes="A2"

for s in wb.worksheets: s.sheet_view.showGridLines=False
wb.move_sheet("1 README",offset=-10)
wb.save("/home/claude/build/qm/KGPL-QM-MasterData-Build.xlsx")
print("saved")
